"""
Multi-Company Irrelevant Lead Audit
Uses Google Gemini AI to analyze mis-categorized leads.
Run: python multi_company_report.py
"""
import re
import gspread
import pandas as pd
from google import genai
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── CONFIG — CHANGE DATES HERE ───────────────────────────────────────────────
FROM_DATE = "2026-04-06"
TO_DATE   = "2026-04-06"
# ─────────────────────────────────────────────────────────────────────────────

GEMINI_API_KEY   = "AIzaSyAFQ5u9zdbf-anhkvb1lfLhQqwnIIw5q8o"
CREDENTIALS_PATH = r"D:\Coding Works\Visual Studio Code\BIPS IRRELEVANT AUDIT\credentials.json.json"
REPORT_PATH      = r"D:\Coding Works\Visual Studio Code\BIPS IRRELEVANT AUDIT\Irrelevant_Audit_Report.xlsx"

COMPANIES = [
    {
        "name":        "BIPS School",
        "description": "School admissions — leads for parents looking to enroll their children",
        "url":         "https://docs.google.com/spreadsheets/d/16uZ_W0EVzxIsBUEzVch5UjD01E_Wzhj5D9tOTD61VnY/edit",
        "worksheet":   "Inquiry",
        "col_date":    2,    # C
        "col_remarks": 7,    # H
        "col_status":  14,   # O
        "col_person":  10,   # K
        "col_source":  None,
    },
    {
        "name":        "Bodyzone Gym",
        "description": "Gym memberships — leads for people looking to join gym or take memberships",
        "url":         "https://docs.google.com/spreadsheets/d/1gBw_HZBnJpFoOARN9yjnHmJleH44FkNgkaSGUsTTqyI/edit",
        "worksheet":   "Inquiry",
        "col_date":    2,    # C
        "col_remarks": 7,    # H
        "col_status":  14,   # O
        "col_person":  9,    # J
        "col_source":  6,    # G
    },
    {
        "name":        "Spa Kora",
        "description": "Spa treatments & memberships — leads for spa services and packages",
        "url":         "https://docs.google.com/spreadsheets/d/1OX7Liv6dg96UxxhScOfYxAk4-h8gxmj6Md4tTwiD1_M/edit",
        "worksheet":   "Inquiry",
        "col_date":    2,    # C
        "col_remarks": 25,   # Z
        "col_status":  10,   # K
        "col_person":  11,   # L
        "col_source":  7,    # H
    },
]

# ── COLORS ────────────────────────────────────────────────────────────────────
DARK_BLUE    = "1F3864"
MID_BLUE     = "2E75B6"
LIGHT_BLUE   = "D6E4F0"
RED          = "C00000"
LIGHT_RED    = "FFE0E0"
GREEN        = "375623"
LIGHT_GREEN  = "E2EFDA"
LIGHT_ORANGE = "FFF2CC"
WHITE        = "FFFFFF"
GREY         = "F2F2F2"
DARK_GREY    = "595959"
COMPANY_COLORS = [
    ("1F3864", "D6E4F0"),  # Blue   — BIPS
    ("7B3F00", "FDEBD0"),  # Brown  — Bodyzone
    ("4A235A", "F5EEF8"),  # Purple — Spa Kora
]
# ─────────────────────────────────────────────────────────────────────────────

def applyc(cell, fill=None, fc="FFFFFF", bold=False, size=10,
           ha="center", va="center", wrap=False, border=False):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=fc, bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if border:
        t = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=t, right=t, top=t, bottom=t)


def load_data(client, company):
    """Pull filtered irrelevant leads from Google Sheet."""
    ws   = client.open_by_url(company["url"]).worksheet(company["worksheet"])
    rows = ws.get_all_values()
    if not rows:
        return pd.DataFrame()

    headers = rows[0]
    seen, uh = {}, []
    for h in headers:
        if h in seen:
            seen[h] += 1; uh.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0; uh.append(h)

    df = pd.DataFrame(rows[1:], columns=uh)

    def col(i): return uh[i] if i is not None and i < len(uh) else None

    c_date    = col(company["col_date"])
    c_remarks = col(company["col_remarks"])
    c_status  = col(company["col_status"])
    c_person  = col(company["col_person"])
    c_source  = col(company["col_source"])

    df["_date"] = pd.to_datetime(df[c_date], dayfirst=True, errors="coerce")
    mask = (
        (df["_date"].dt.date >= pd.Timestamp(FROM_DATE).date()) &
        (df["_date"].dt.date <= pd.Timestamp(TO_DATE).date()) &
        (df[c_status].str.strip().str.lower() == "irrelevant")
    )
    f = df[mask].copy().reset_index(drop=True)
    f["_date_str"] = f[c_date]
    f["_remarks"]  = f[c_remarks].fillna("").str.strip() if c_remarks else ""
    f["_person"]   = f[c_person].fillna("").str.strip()  if c_person  else ""
    f["_source"]   = f[c_source].fillna("").str.strip()  if c_source  else ""
    return f[["_date_str", "_remarks", "_person", "_source"]]


def analyze_with_gemini(company, df):
    """Send all leads to Gemini in one call, get verdicts back."""
    if df.empty:
        return []

    leads_text = ""
    for i, row in df.iterrows():
        leads_text += f"LEAD {i+1}:\nRemarks: {row['_remarks'] or 'No remarks provided'}\n\n"

    prompt = f"""You are a sales audit expert for: {company['description']}

All leads below were marked IRRELEVANT by the sales caller. Analyze each remark and decide if the lead was MIS-CATEGORIZED (genuine potential >=25% to become a customer).

{leads_text}

Respond for EVERY lead in EXACTLY this format — no skipping:

LEAD 1:
POTENTIAL: [0% / 10% / 25% / 35% / 50%]
VERDICT: [Correctly Irrelevant / MIS-CATEGORIZED]
REASON: [1 sentence]
ACTION: [1 sentence recommended next step, or None]

LEAD 2:
...and so on for all {len(df)} leads.

Rules:
- MIS-CATEGORIZED only if potential >= 25%
- Genuine interest, specific questions about service, real addressable need = MIS-CATEGORIZED
- Wrong location/distance alone = Correctly Irrelevant
- Wrong number, accidental click, no interest = Correctly Irrelevant"""

    client   = genai.Client(api_key=GEMINI_API_KEY)
    response = client.models.generate_content(model="gemini-2.0-flash", contents=prompt)
    result   = response.text

    # Parse response
    blocks = re.split(r'LEAD \d+:', result)
    blocks = [b.strip() for b in blocks if b.strip()]

    results = []
    for i, row in df.iterrows():
        block  = blocks[i] if i < len(blocks) else ""
        parsed = {"potential": "—", "verdict": "Correctly Irrelevant",
                  "reason": "—", "action": "None"}
        for line in block.split("\n"):
            line = line.strip()
            if line.startswith("POTENTIAL:"):
                parsed["potential"] = line.replace("POTENTIAL:", "").strip()
            elif line.startswith("VERDICT:"):
                parsed["verdict"] = line.replace("VERDICT:", "").strip()
            elif line.startswith("REASON:"):
                parsed["reason"] = line.replace("REASON:", "").strip()
            elif line.startswith("ACTION:"):
                parsed["action"] = line.replace("ACTION:", "").strip()
        results.append({
            "date":      row["_date_str"],
            "person":    row["_person"],
            "source":    row["_source"],
            "remarks":   row["_remarks"],
            "potential": parsed["potential"],
            "verdict":   parsed["verdict"],
            "reason":    parsed["reason"],
            "action":    parsed["action"],
        })
    return results


def build_company_sheets(wb, company, results, color_pair):
    dark, light = color_pair
    total   = len(results)
    flagged = [r for r in results if "MIS" in r.get("verdict", "")]
    acc     = round((total - len(flagged)) / total * 100, 1) if total else 0
    date_label = FROM_DATE if FROM_DATE == TO_DATE else f"{FROM_DATE} to {TO_DATE}"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet(f"{company['name']} Summary")
    ws.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E","F","G"], [3,20,20,16,16,16,20]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:G1")
    ws["B1"] = f"IRRELEVANT LEAD AUDIT — {company['name'].upper()}"
    applyc(ws["B1"], fill=dark, bold=True, size=16)
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B2:G2")
    ws["B2"] = f"Date: {date_label}  |  {company['description']}  |  Prepared: {datetime.today().strftime('%d %b %Y')}"
    applyc(ws["B2"], fill=MID_BLUE, size=10)
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 12

    # KPI cards
    kpis = [
        ("B4:C5", f"Total Leads\nReviewed\n{total}",              dark),
        ("D4:E5", f"Correctly\nIrrelevant\n{total-len(flagged)}", MID_BLUE),
        ("F4:F5", f"Mis-Cat.\nFound\n{len(flagged)}",             RED),
        ("G4:G5", f"Accuracy\n\n{acc}%",                          GREEN),
    ]
    for merge, text, color in kpis:
        ws.merge_cells(merge)
        ws[merge.split(":")[0]] = text
        applyc(ws[merge.split(":")[0]], fill=color, bold=True, size=12, wrap=True)
    ws.row_dimensions[4].height = 50
    ws.row_dimensions[5].height = 50
    ws.row_dimensions[6].height = 12

    # Flagged leads
    ws.merge_cells("B7:G7")
    ws["B7"] = "  MIS-CATEGORIZED LEADS — ACTION REQUIRED"
    applyc(ws["B7"], fill=RED, bold=True, size=12, ha="left")
    ws.row_dimensions[7].height = 26

    if flagged:
        for ci, (h, col, w) in enumerate(zip(
            ["#", "Date", "Salesperson", "Potential", "Why Flagged", "Recommended Action"],
            ["B","C","D","E","F","G"],
            [4, 16, 16, 12, 38, 38]
        ), 0):
            ws.column_dimensions[col].width = w
            c = ws.cell(row=8, column=ci+2, value=h)
            applyc(c, fill=DARK_GREY, bold=True, size=10)
        ws.row_dimensions[8].height = 20

        for ri, lead in enumerate(flagged, 9):
            for ci, val in enumerate(
                [ri-8, lead["date"], lead["person"], lead["potential"],
                 lead["reason"], lead["action"]], 2
            ):
                c = ws.cell(row=ri, column=ci, value=val)
                applyc(c, fill=LIGHT_RED, fc=RED, ha="left", size=9,
                       border=True, wrap=True, bold=(ci == 6))
            ws.row_dimensions[ri].height = 45
    else:
        ws.merge_cells("B8:G8")
        ws["B8"] = "No mis-categorized leads found for this period."
        applyc(ws["B8"], fill=LIGHT_GREEN, fc=GREEN, ha="left", size=10)
        ws.row_dimensions[8].height = 22

    # ── All Leads sheet ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet(f"{company['name']} All Leads")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"ALL IRRELEVANT LEADS — {company['name'].upper()} — {date_label}"
    applyc(ws2["A1"], fill=dark, bold=True, size=13)
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells("A2:G2")
    ws2["A2"] = f"Total: {total}  |  Mis-categorized: {len(flagged)}  |  Accuracy: {acc}%"
    applyc(ws2["A2"], fill=MID_BLUE, size=10)
    ws2.row_dimensions[2].height = 18
    ws2.row_dimensions[3].height = 8

    hdrs   = ["#", "Date & Time", "Salesperson", "Lead Source", "Remarks", "Verdict", "Potential"]
    widths = [5, 20, 16, 16, 52, 24, 12]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
        c = ws2.cell(row=4, column=ci, value=h)
        applyc(c, fill=dark, bold=True, size=10)
    ws2.row_dimensions[4].height = 22

    for ri, lead in enumerate(results, 5):
        is_flag = "MIS" in lead.get("verdict", "")
        bg = LIGHT_RED if is_flag else (GREY if ri % 2 == 0 else WHITE)
        fc = RED if is_flag else "000000"
        vals = [ri-4, lead["date"], lead["person"], lead["source"],
                lead["remarks"].replace("\n"," "), lead["verdict"], lead["potential"]]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            applyc(c, fill=bg, fc=fc, ha="left", size=9, border=True, wrap=(ci==5))
            if is_flag:
                c.font = Font(color=fc, size=9, bold=True, name="Calibri")
        ws2.row_dimensions[ri].height = 38 if is_flag else 28


def build_overview(wb, all_data):
    ws = wb.active
    ws.title = "Combined Overview"
    ws.sheet_view.showGridLines = False
    for col, w in zip(["A","B","C","D","E","F","G","H"], [3,22,18,14,14,14,14,22]):
        ws.column_dimensions[col].width = w

    date_label = FROM_DATE if FROM_DATE == TO_DATE else f"{FROM_DATE} to {TO_DATE}"

    ws.merge_cells("B1:H1")
    ws["B1"] = "MULTI-COMPANY IRRELEVANT LEAD AUDIT REPORT"
    applyc(ws["B1"], fill=DARK_BLUE, bold=True, size=18)
    ws.row_dimensions[1].height = 42

    ws.merge_cells("B2:H2")
    ws["B2"] = f"Date: {date_label}   |   Prepared: {datetime.today().strftime('%d %B %Y')}"
    applyc(ws["B2"], fill=MID_BLUE, size=11)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 14

    ws.merge_cells("B4:H4")
    ws["B4"] = "  COMPANY-WISE SUMMARY"
    applyc(ws["B4"], fill=DARK_BLUE, bold=True, size=12, ha="left")
    ws.row_dimensions[4].height = 26

    for ci, h in enumerate(
        ["Company", "Description", "Total Leads", "Correctly Irrelevant",
         "Mis-Categorized", "Accuracy %", "Status"], 2
    ):
        c = ws.cell(row=5, column=ci, value=h)
        applyc(c, fill=DARK_GREY, bold=True, size=10)
    ws.row_dimensions[5].height = 20

    grand_total = grand_flag = 0
    for ri, (company, results, color_pair) in enumerate(all_data, 6):
        _, light = color_pair
        total   = len(results)
        flagged = sum(1 for r in results if "MIS" in r.get("verdict", ""))
        correct = total - flagged
        acc     = round(correct / total * 100, 1) if total else 0
        grand_total += total; grand_flag += flagged
        status = "Action Needed" if flagged > 0 else ("No Leads" if total == 0 else "All Clear")
        sc = LIGHT_RED if flagged > 0 else (GREY if total == 0 else LIGHT_GREEN)
        sfc = RED if flagged > 0 else (DARK_GREY if total == 0 else GREEN)
        vals = [company["name"], company["description"][:40], total, correct, flagged, f"{acc}%", status]
        for ci, val in enumerate(vals, 2):
            c = ws.cell(row=ri, column=ci, value=val)
            bg = sc if ci == 8 else (light if ri % 2 == 0 else WHITE)
            fc = sfc if ci == 8 else "000000"
            applyc(c, fill=bg, fc=fc, ha="left" if ci == 2 else "center",
                   size=10, border=True)
        ws.row_dimensions[ri].height = 22

    gr = 6 + len(all_data)
    grand_acc = round((grand_total - grand_flag) / grand_total * 100, 1) if grand_total else 0
    for ci, val in enumerate(
        ["TOTAL", "", grand_total, grand_total-grand_flag, grand_flag, f"{grand_acc}%", ""], 2
    ):
        c = ws.cell(row=gr, column=ci, value=val)
        applyc(c, fill=DARK_BLUE, bold=True, size=10, border=True)
    ws.row_dimensions[gr].height = 24

    # All flagged leads combined
    ws.row_dimensions[gr+1].height = 16
    ws.merge_cells(f"B{gr+2}:H{gr+2}")
    ws[f"B{gr+2}"] = "  ALL MIS-CATEGORIZED LEADS ACROSS COMPANIES"
    applyc(ws[f"B{gr+2}"], fill=RED, bold=True, size=12, ha="left")
    ws.row_dimensions[gr+2].height = 26

    for ci, h in enumerate(
        ["Company", "Date", "Salesperson", "Potential", "Remarks", "Reason", "Action"], 2
    ):
        c = ws.cell(row=gr+3, column=ci, value=h)
        applyc(c, fill=DARK_GREY, bold=True, size=10)
    ws.row_dimensions[gr+3].height = 20

    crow = gr + 4
    any_flag = False
    for company, results, _ in all_data:
        for lead in results:
            if "MIS" in lead.get("verdict", ""):
                any_flag = True
                vals = [company["name"], lead["date"], lead["person"],
                        lead["potential"], lead["remarks"][:80].replace("\n"," "),
                        lead["reason"], lead["action"]]
                for ci, val in enumerate(vals, 2):
                    c = ws.cell(row=crow, column=ci, value=val)
                    applyc(c, fill=LIGHT_RED, fc=RED, ha="left", size=9,
                           border=True, wrap=True)
                ws.row_dimensions[crow].height = 45
                crow += 1

    if not any_flag:
        ws.merge_cells(f"B{gr+4}:H{gr+4}")
        ws[f"B{gr+4}"] = "No mis-categorized leads found across any company."
        applyc(ws[f"B{gr+4}"], fill=LIGHT_GREEN, fc=GREEN, ha="left", size=10)
        ws.row_dimensions[gr+4].height = 22


def main():
    print(f"Irrelevant Lead Audit | {FROM_DATE} to {TO_DATE}")
    print("=" * 50)

    gspread_client = gspread.service_account(filename=CREDENTIALS_PATH)

    wb       = Workbook()
    all_data = []

    for i, company in enumerate(COMPANIES):
        print(f"\n[{i+1}/3] {company['name']}")
        try:
            print(f"  Loading sheet...", end=" ")
            df = load_data(gspread_client, company)
            print(f"{len(df)} Irrelevant leads found.")

            if df.empty:
                all_data.append((company, [], COMPANY_COLORS[i]))
                build_company_sheets(wb, company, [], COMPANY_COLORS[i])
                continue

            print(f"  Analyzing with Gemini AI...", end=" ", flush=True)
            results = analyze_with_gemini(company, df)
            flagged = sum(1 for r in results if "MIS" in r.get("verdict",""))
            print(f"Done. {flagged} mis-categorized found.")

            build_company_sheets(wb, company, results, COMPANY_COLORS[i])
            all_data.append((company, results, COMPANY_COLORS[i]))
        except Exception as e:
            print(f"  ERROR: {e}")
            all_data.append((company, [], COMPANY_COLORS[i]))

    print("\nBuilding report...")
    build_overview(wb, all_data)
    wb.move_sheet("Combined Overview", offset=-len(wb.sheetnames)+1)
    wb.save(REPORT_PATH)
    print(f"Report saved: {REPORT_PATH}")

    print("\n" + "=" * 50)
    for company, results, _ in all_data:
        flagged = sum(1 for r in results if "MIS" in r.get("verdict",""))
        print(f"  {company['name']}: {len(results)} leads | {flagged} mis-categorized")


if __name__ == "__main__":
    main()
