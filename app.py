"""
CRM Analysis AI Agent — Web App
Powered by Groq AI + Google Sheets
Run: python app.py  →  open http://localhost:5000
"""
import re, io, json, os, tempfile
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from groq import Groq
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── CONFIG ────────────────────────────────────────────────────────────────────
GROQ_API_KEY     = os.environ.get("GROQ_API_KEY", "")
CREDENTIALS_PATH = r"D:\Coding Works\Visual Studio Code\BIPS IRRELEVANT AUDIT\credentials.json.json"

def get_gspread_client():
    """Connect to Google Sheets — supports both local file and cloud env var."""
    scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if creds_json:
        creds_dict = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        return gspread.authorize(creds)
    elif os.path.exists(CREDENTIALS_PATH):
        return gspread.service_account(filename=CREDENTIALS_PATH)
    else:
        raise Exception("GOOGLE_CREDENTIALS_JSON environment variable is not set in Railway. Go to Railway → Variables and add it.")

COMPANIES = {
    "bips": {
        "name":        "BIPS School",
        "description": "School that handles admissions. Leads are parents inquiring about enrolling their children.",
        "url":         "https://docs.google.com/spreadsheets/d/16uZ_W0EVzxIsBUEzVch5UjD01E_Wzhj5D9tOTD61VnY/edit",
        "worksheet":   "Inquiry",
        "col_date":    2,    # C
        "col_remarks": 7,    # H
        "col_status":  14,   # O
        "col_person":  10,   # K
        "col_source":  None,
    },
    "bodyzone": {
        "name":        "Bodyzone Gym",
        "description": "Gym that sells memberships. Leads are people inquiring about joining the gym.",
        "url":         "https://docs.google.com/spreadsheets/d/1gBw_HZBnJpFoOARN9yjnHmJleH44FkNgkaSGUsTTqyI/edit",
        "worksheet":   "Inquiry",
        "col_date":    2,    # C
        "col_remarks": 7,    # H
        "col_status":  14,   # O
        "col_person":  9,    # J
        "col_source":  6,    # G
    },
}

groq_client    = Groq(api_key=GROQ_API_KEY)
gspread_client = None

# ── HELPERS ───────────────────────────────────────────────────────────────────
def get_gspread():
    global gspread_client
    if gspread_client is None:
        gspread_client = get_gspread_client()
    return gspread_client

def load_sheet(company_key, from_date, to_date, status_filter):
    co   = COMPANIES[company_key]
    ws   = get_gspread().open_by_url(co["url"]).worksheet(co["worksheet"])
    # Only fetch columns A through BZ to prevent hanging on accidentally widened sheets
    rows = ws.get("A:BZ")
    if not rows or len(rows) < 2:
        return pd.DataFrame()

    # Build unique headers (handle duplicate column names)
    raw_headers = rows[0]
    seen, uh = {}, []
    for h in raw_headers:
        if h in seen:
            seen[h] += 1
            uh.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            uh.append(h)

    # Normalize all data rows to match header length (some rows may be shorter/longer)
    header_len = len(uh)
    norm_rows = []
    for row in rows[1:]:
        if len(row) < header_len:
            row = row + [''] * (header_len - len(row))
        elif len(row) > header_len:
            row = row[:header_len]
        norm_rows.append(row)

    df = pd.DataFrame(norm_rows, columns=uh)

    # Get column names by index (use original index, not de-duped name)
    def col_name(idx):
        return uh[idx] if idx is not None and idx < len(uh) else None

    c_date    = col_name(co["col_date"])
    c_remarks = col_name(co["col_remarks"])
    c_status  = col_name(co["col_status"])
    c_person  = col_name(co["col_person"])
    c_source  = col_name(co["col_source"])

    # Parse date (BIPS uses DD/MM/YYYY, Bodyzone/SpaKora use M/D/YYYY)
    is_dayfirst = (company_key == "bips")
    df["_date"] = pd.to_datetime(df[c_date], format="mixed", dayfirst=is_dayfirst, errors="coerce")

    # Build filter
    mask = (
        (df["_date"].dt.date >= pd.Timestamp(from_date).date()) &
        (df["_date"].dt.date <= pd.Timestamp(to_date).date())
    )
    if status_filter and status_filter != "all":
        mask &= df[c_status].str.strip().str.lower() == status_filter.lower()

    f = df[mask].copy().reset_index(drop=True)
    if f.empty:
        return pd.DataFrame()

    f["_date_str"] = f[c_date]
    f["_remarks"]  = f[c_remarks].fillna("").str.strip() if c_remarks else ""
    f["_person"]   = f[c_person].fillna("").str.strip()  if c_person  else ""
    f["_source"]   = f[c_source].fillna("").str.strip()  if c_source  else ""
    f["_status"]   = f[c_status].fillna("").str.strip()  if c_status  else ""
    return f[["_date_str", "_remarks", "_person", "_source", "_status"]]


def analyze_with_groq(company_key, df, status_filter):
    co = COMPANIES[company_key]
    if df.empty:
        return []

    leads_text = ""
    for i, row in df.reset_index(drop=True).iterrows():
        leads_text += f"LEAD {i+1}:\nCurrent Status: {row['_status']}\nRemarks: {row['_remarks'] or 'No remarks provided'}\n\n"

    if status_filter == "irrelevant":
        task = """The telecaller marked these leads as IRRELEVANT.
Your job: Check if the telecaller was WRONG — i.e., does the lead actually have >=25% chance of becoming a paying customer?
- If YES (potential >=25%) → MIS-CATEGORIZED (wrongly marked irrelevant)
- If NO (potential <25%)  → Correctly Irrelevant"""
    elif status_filter == "relevant":
        task = """The telecaller marked these leads as RELEVANT.
Your job: Check if the telecaller was WRONG — i.e., is the lead actually NOT a genuine prospect?
- If the person has no real interest, wrong number, not the target customer → MIS-CATEGORIZED (wrongly marked relevant)
- If the person genuinely showed interest or has a real need → Correctly Relevant"""
    else:
        task = """Analyze each lead based on its current status and remarks.
- If IRRELEVANT lead has >=25% potential → MIS-CATEGORIZED
- If RELEVANT lead has no genuine interest → MIS-CATEGORIZED
- Otherwise → Correctly Categorized"""

    prompt = f"""You are a senior sales audit expert for: {co['description']}

{task}

{leads_text}

Respond for EVERY lead in EXACTLY this format — do not skip any:

LEAD 1:
POTENTIAL: [0% / 10% / 25% / 35% / 50%]
VERDICT: [Correctly Irrelevant / Correctly Relevant / MIS-CATEGORIZED]
REASON: [1 sentence explanation]
ACTION: [1 sentence recommended next step, or None]

Rules for MIS-CATEGORIZED:
- Irrelevant leads: genuine interest, asked about fees/facilities/timings, real need = MIS-CATEGORIZED
- Relevant leads: wrong number, no interest, not target audience = MIS-CATEGORIZED
- Wrong location/distance alone for irrelevant = Correctly Irrelevant
- Accidental click, clearly no interest = Correctly Irrelevant"""

    resp = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    result = resp.choices[0].message.content
    blocks = [b.strip() for b in re.split(r'LEAD \d+:', result) if b.strip()]

    results = []
    for i, row in df.reset_index(drop=True).iterrows():
        block  = blocks[i] if i < len(blocks) else ""
        parsed = {"potential": "—", "verdict": "—", "reason": "—", "action": "None"}
        for line in block.split("\n"):
            l = line.strip()
            if l.startswith("POTENTIAL:"): parsed["potential"] = l.replace("POTENTIAL:", "").strip()
            elif l.startswith("VERDICT:"):  parsed["verdict"]   = l.replace("VERDICT:", "").strip()
            elif l.startswith("REASON:"):   parsed["reason"]    = l.replace("REASON:", "").strip()
            elif l.startswith("ACTION:"):   parsed["action"]    = l.replace("ACTION:", "").strip()
        results.append({
            "date":      row["_date_str"],
            "person":    row["_person"],
            "source":    row["_source"],
            "remarks":   row["_remarks"],
            "status":    row["_status"],
            "potential": parsed["potential"],
            "verdict":   parsed["verdict"],
            "reason":    parsed["reason"],
            "action":    parsed["action"],
            "flagged":   "MIS" in parsed["verdict"],
        })
    return results


def parse_command_with_groq(command):
    """Use Groq to determine if command is an analyze request or a general question."""
    today = datetime.today().strftime("%Y-%m-%d")
    prompt = f"""You are a CRM assistant. Analyze this user command. Today's date is {today}.

Command: "{command}"

First decide: is this an ANALYZE command (user wants to analyze/audit/check leads by status) or a GENERAL QUESTION (user is asking anything else about the data)?

Return ONLY a JSON object:
- If ANALYZE: {{"type":"analyze","company":"bips|bodyzone|spakora","from_date":"YYYY-MM-DD","to_date":"YYYY-MM-DD","status":"irrelevant|relevant|all"}}
- If GENERAL: {{"type":"question","company":"bips|bodyzone|spakora","from_date":"YYYY-MM-DD","to_date":"YYYY-MM-DD","question":"{command}"}}

Default company: bips. Default status: irrelevant. For general questions, set wide date range if no date mentioned (e.g. 2026-04-01 to {today}).
Return ONLY valid JSON."""

    resp = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
    )
    text = resp.choices[0].message.content.strip()
    match = re.search(r'\{.*?\}', text, re.DOTALL)
    if match:
        return json.loads(match.group())
    return None


def answer_general_question(company_key, from_date, to_date, question):
    """Load full sheet data and answer any question using Groq."""
    co  = COMPANIES[company_key]
    ws  = get_gspread().open_by_url(co["url"]).worksheet(co["worksheet"])
    rows = ws.get("A:BZ")
    if not rows or len(rows) < 2:
        return "No data found for this company."

    # Build text summary of data (limit rows to avoid token overflow)
    header = rows[0]
    data_rows = rows[1:]

    # Filter by date if possible
    try:
        date_col = co["col_date"]
        is_dayfirst = (company_key == "bips")
        filtered = []
        for row in data_rows:
            if len(row) > date_col and row[date_col]:
                try:
                    d = pd.to_datetime(row[date_col], dayfirst=is_dayfirst, errors="coerce")
                    if pd.notna(d):
                        if pd.Timestamp(from_date).date() <= d.date() <= pd.Timestamp(to_date).date():
                            filtered.append(row)
                except:
                    pass
        data_rows = filtered if filtered else data_rows
    except:
        pass

    # Cap at 200 rows to avoid token overflow
    data_rows = data_rows[:200]

    # Format as readable table
    table_lines = [" | ".join(str(v) for v in header)]
    for row in data_rows:
        padded = row + [''] * (len(header) - len(row))
        table_lines.append(" | ".join(str(v) for v in padded[:len(header)]))
    table_text = "\n".join(table_lines)

    prompt = f"""You are a CRM data analyst for {co['name']}.
Here is the sheet data (date range {from_date} to {to_date}):

{table_text}

Answer this question clearly and concisely:
{question}

Give a direct, helpful answer with numbers/stats where relevant."""

    resp = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        max_tokens=1000,
    )
    return resp.choices[0].message.content.strip()


def build_excel(company_key, results, from_date, to_date):
    co      = COMPANIES[company_key]
    wb      = Workbook()
    ws      = wb.active
    ws.title = "Audit Report"
    ws.sheet_view.showGridLines = False

    DARK  = "1F3864"; MID   = "2E75B6"; RED   = "C00000"
    LRED  = "FFE0E0"; GREY  = "F2F2F2"; WHITE = "FFFFFF"
    GREEN = "375623"; LGRN  = "E2EFDA"; DGREY = "595959"

    def s(cell, fill=None, fc="FFFFFF", bold=False, size=10,
          ha="center", wrap=False, border=False):
        if fill: cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color=fc, bold=bold, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
        if border:
            t = Side(style="thin", color="BFBFBF")
            cell.border = Border(left=t, right=t, top=t, bottom=t)

    date_label = from_date if from_date == to_date else f"{from_date} to {to_date}"
    total   = len(results)
    flagged = [r for r in results if r.get("flagged")]
    acc     = round((total - len(flagged)) / total * 100, 1) if total else 0

    for c, w in zip(["A","B","C","D","E","F","G","H","I"],
                    [5, 20, 18, 14, 48, 20, 12, 26, 30]):
        ws.column_dimensions[c].width = w

    ws.merge_cells("A1:I1")
    ws["A1"] = f"LEAD AUDIT REPORT — {co['name'].upper()} — {date_label}"
    s(ws["A1"], fill=DARK, bold=True, size=15)
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Total: {total}  |  Mis-categorized: {len(flagged)}  |  Accuracy: {acc}%  |  Prepared: {datetime.today().strftime('%d %b %Y')}"
    s(ws["A2"], fill=MID, size=10)
    ws.row_dimensions[2].height = 20

    kpis = [("A3:B4",f"Leads\nReviewed\n{total}",DARK),
            ("C3:E4",f"Correctly Categorized\n\n{total-len(flagged)}",MID),
            ("F3:G4",f"Mis-Categorized\n\n{len(flagged)}",RED),
            ("H3:I4",f"Accuracy\n\n{acc}%",GREEN)]
    for merge, text, color in kpis:
        ws.merge_cells(merge)
        ws[merge.split(":")[0]] = text
        s(ws[merge.split(":")[0]], fill=color, bold=True, size=12, wrap=True)
    ws.row_dimensions[3].height = 45
    ws.row_dimensions[4].height = 45
    ws.row_dimensions[5].height = 10

    hdrs = ["#","Date & Time","Salesperson","Status","Remarks","Verdict","Potential","Reason","Action"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=6, column=ci, value=h)
        s(c, fill=DGREY, bold=True, size=10)
    ws.row_dimensions[6].height = 22

    for ri, lead in enumerate(results, 7):
        is_f = lead.get("flagged", False)
        bg = LRED if is_f else (GREY if ri % 2 == 0 else WHITE)
        fc = RED  if is_f else "000000"
        vals = [ri-6, lead["date"], lead["person"], lead["status"],
                lead["remarks"].replace("\n"," "), lead["verdict"],
                lead["potential"], lead["reason"], lead["action"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            s(c, fill=bg, fc=fc, ha="left", size=9, border=True, wrap=(ci in [5,8,9]))
            if is_f: c.font = Font(color=fc, size=9, bold=True, name="Calibri")
        ws.row_dimensions[ri].height = 42 if is_f else 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/debug")
def debug():
    creds = os.environ.get("GOOGLE_CREDENTIALS_JSON", "NOT SET")
    groq  = os.environ.get("GROQ_API_KEY", "NOT SET")
    return {
        "GOOGLE_CREDENTIALS_JSON": f"SET ({len(creds)} chars)" if creds != "NOT SET" else "NOT SET",
        "GROQ_API_KEY": "SET" if groq != "NOT SET" else "NOT SET"
    }

@app.route("/analyze", methods=["POST"])
def analyze():
    data        = request.json
    company_key = data.get("company", "bips")
    from_date   = data.get("from_date", "")
    to_date     = data.get("to_date", "")
    status      = data.get("status", "irrelevant")

    if not from_date or not to_date:
        return jsonify({"error": "Please select a date range."}), 400
    if company_key not in COMPANIES:
        return jsonify({"error": "Invalid company selected."}), 400

    try:
        df = load_sheet(company_key, from_date, to_date, status)
        if df.empty:
            return jsonify({
                "results": [], "total": 0, "flagged": 0, "accuracy": 0,
                "company": COMPANIES[company_key]["name"],
                "from_date": from_date,
                "to_date": to_date,
                "status": status,
                "message": f"No '{status}' leads found for the selected date range."
            })

        results = analyze_with_groq(company_key, df, status)
        flagged = sum(1 for r in results if r.get("flagged"))
        return jsonify({
            "results":   results,
            "total":     len(results),
            "flagged":   flagged,
            "accuracy":  round((len(results)-flagged)/len(results)*100,1),
            "company":   COMPANIES[company_key]["name"],
            "from_date": from_date,
            "to_date":   to_date,
            "status":    status,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/command", methods=["POST"])
def command():
    """Parse a natural language command — handles both analyze and general questions."""
    text = request.json.get("text", "").strip()
    if not text:
        return jsonify({"error": "Please type a command."}), 400
    try:
        params = parse_command_with_groq(text)
        if not params:
            return jsonify({"error": "Could not understand the command."}), 400

        company_key = params.get("company", "bips")
        from_date   = params.get("from_date", "")
        to_date     = params.get("to_date", "")
        if company_key not in COMPANIES:
            company_key = "bips"

        # General question — answer directly from sheet data
        if params.get("type") == "question":
            answer = answer_general_question(company_key, from_date, to_date, params.get("question", text))
            return jsonify({"type": "answer", "answer": answer, "company": COMPANIES[company_key]["name"], "from_date": from_date, "to_date": to_date})

        # Analyze command
        status = params.get("status", "irrelevant")
        df = load_sheet(company_key, from_date, to_date, status)
        if df.empty:
            return jsonify({
                "results": [], "total": 0, "flagged": 0, "accuracy": 0,
                "company": COMPANIES[company_key]["name"],
                "from_date": from_date, "to_date": to_date, "status": status,
                "message": f"No '{status}' leads found for the selected period.",
            })

        results = analyze_with_groq(company_key, df, status)
        flagged = sum(1 for r in results if r.get("flagged"))
        return jsonify({
            "results":   results,
            "total":     len(results),
            "flagged":   flagged,
            "accuracy":  round((len(results)-flagged)/len(results)*100,1),
            "company":   COMPANIES[company_key]["name"],
            "from_date": from_date,
            "to_date":   to_date,
            "status":    status,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/download", methods=["POST"])
def download():
    data        = request.json
    company_key = data.get("company", "bips")
    from_date   = data.get("from_date", "")
    to_date     = data.get("to_date", "")
    results     = data.get("results", [])

    buf      = build_excel(company_key, results, from_date, to_date)
    co       = COMPANIES[company_key]
    filename = f"Lead_Audit_{co['name'].replace(' ','_')}_{from_date}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"CRM Analysis AI Agent running at http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port)
