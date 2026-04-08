import gspread
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime

# ── CONFIG ────────────────────────────────────────────────────────────────────
CREDENTIALS_PATH = r"D:\Coding Works\Visual Studio Code\BIPS IRRELEVANT AUDIT\credentials.json.json"
SHEET_URL        = "https://docs.google.com/spreadsheets/d/16uZ_W0EVzxIsBUEzVch5UjD01E_Wzhj5D9tOTD61VnY/edit"
WORKSHEET_NAME   = "Inquiry"
REPORT_PATH      = r"D:\Coding Works\Visual Studio Code\BIPS IRRELEVANT AUDIT\Irrelevant_Audit_Report.xlsx"
FROM_DATE        = "2026-04-05"
TO_DATE          = "2026-04-07"
# ─────────────────────────────────────────────────────────────────────────────

# ── COLORS ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
RED         = "C00000"
ORANGE      = "E26B0A"
GREEN       = "375623"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED   = "FFE0E0"
LIGHT_ORANGE= "FFF2CC"
YELLOW      = "FFC000"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"
DARK_GREY   = "595959"
# ─────────────────────────────────────────────────────────────────────────────

def style(fill=None, font_color=WHITE, bold=False, size=11,
          h_align="center", v_align="center", wrap=False, border=False):
    s = {}
    if fill:
        s["fill"] = PatternFill("solid", fgColor=fill)
    s["font"] = Font(color=font_color, bold=bold, size=size, name="Calibri")
    s["alignment"] = Alignment(horizontal=h_align, vertical=v_align,
                                wrap_text=wrap)
    if border:
        thin = Side(style="thin", color="BFBFBF")
        s["border"] = Border(left=thin, right=thin, top=thin, bottom=thin)
    return s

def apply(cell, **kwargs):
    s = style(**kwargs)
    for k, v in s.items():
        setattr(cell, k, v)

def thick_border(ws, min_row, max_row, min_col, max_col, color="1F3864"):
    med = Side(style="medium", color=color)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            t = cell.border.top   if cell.row != min_row else med
            b = cell.border.bottom if cell.row != max_row else med
            l = cell.border.left  if cell.col_idx != min_col else med
            r = cell.border.right if cell.col_idx != max_col else med
            cell.border = Border(left=l, right=r, top=t, bottom=b)

# ─────────────────────────────────────────────────────────────────────────────
print("Connecting to Google Sheets...")
client    = gspread.service_account(filename=CREDENTIALS_PATH)
sheet     = client.open_by_url(SHEET_URL)
worksheet = sheet.worksheet(WORKSHEET_NAME)
rows      = worksheet.get_all_values()
headers   = rows[0]

seen, unique_headers = {}, []
for h in headers:
    if h in seen:
        seen[h] += 1
        unique_headers.append(f"{h}_{seen[h]}")
    else:
        seen[h] = 0
        unique_headers.append(h)

df = pd.DataFrame(rows[1:], columns=unique_headers)

col_c = unique_headers[2]   # Inquiry Date & Time
col_h = unique_headers[7]   # Update Remarks
col_o = unique_headers[14]  # Inquiry Status
col_k = unique_headers[10]  # Sales Person Email

df["_date"] = pd.to_datetime(df[col_c], dayfirst=True, errors="coerce")

mask = (
    (df["_date"].dt.date >= pd.Timestamp(FROM_DATE).date()) &
    (df["_date"].dt.date <= pd.Timestamp(TO_DATE).date()) &
    (df[col_o].str.strip().str.lower() == "irrelevant")
)
fdf = df[mask].copy().reset_index(drop=True)

# Categorize each lead
reasons = [
    "Wrong location / too far",
    "Wrong location / too far",
    "Wrong location / too far",
    "Wrong location / too far",
    "Dead lead (no contact)",
    "Wrong location / too far",
    "Wrong location / too far",
    "Wrong location / too far",
    "Wrong location / too far",
    "Wrong location / unmarried",
    "Wrong location / too far",
    "Wrong location / too far",
    "Wrong location / too far",
    "POTENTIAL – Wants branch near Zirakpur",   # Lead 14 FLAGGED
    "Inquired by mistake",
    "Inquired by mistake",
    "Wrong location / too far",
    "Wrong location / too far",
    "POTENTIAL – Asking about Hostel facility", # Lead 19 FLAGGED
    "Non-admission inquiry (Conference)",
    "Wrong location + affordability issue",
    "Inquired by mistake",
    "Non-admission inquiry (Bus incharge)",
    "Inquired by mistake",
    "Non-admission inquiry (Competition event)",
    "Wrong location / too far",
    "Wrong number (dialed school by mistake)",
    "Existing parent (not new inquiry)",
    "Wrong location / too far",
]

potentials = [
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "MIS-CATEGORIZED (~35-40%)",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "MIS-CATEGORIZED (~25-30%)",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
    "Correctly Irrelevant",
]

fdf["Reason"]    = reasons
fdf["Verdict"]   = potentials

# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 3
for col in ["B","C","D","E","F","G","H"]:
    ws1.column_dimensions[col].width = 18

# Title block
ws1.merge_cells("B1:H1")
ws1["B1"] = "IRRELEVANT LEAD AUDIT REPORT"
apply(ws1["B1"], fill=DARK_BLUE, bold=True, size=18)
ws1.row_dimensions[1].height = 40

ws1.merge_cells("B2:H2")
ws1["B2"] = f"Period: 5th April – 7th April 2026   |   Worksheet: {WORKSHEET_NAME}   |   Prepared: {datetime.today().strftime('%d %b %Y')}"
apply(ws1["B2"], fill=MID_BLUE, size=10)
ws1.row_dimensions[2].height = 22

ws1.row_dimensions[3].height = 14

# KPI Cards row
ws1.row_dimensions[4].height = 20
ws1.row_dimensions[5].height = 45
ws1.row_dimensions[6].height = 30
ws1.row_dimensions[7].height = 20

kpis = [
    ("B5:C6", "Total Irrelevant\nLeads Reviewed", "29", DARK_BLUE),
    ("D5:E6", "Correctly\nIrrelevant", "27", MID_BLUE),
    ("F5:G6", "Mis-Categorized\nLeads Found", "2", RED),
    ("H5:H6", "Accuracy\nRate", "93%", "375623"),
]
for merge, label, val, color in kpis:
    ws1.merge_cells(merge)
    first_cell = merge.split(":")[0]
    ws1[first_cell] = f"{label}\n{val}"
    apply(ws1[first_cell], fill=color, bold=True, size=13, wrap=True)

ws1.row_dimensions[8].height = 18

# Section: Key Findings
ws1.merge_cells("B9:H9")
ws1["B9"] = "  KEY FINDINGS"
apply(ws1["B9"], fill=MID_BLUE, bold=True, size=12, h_align="left")
ws1.row_dimensions[9].height = 26

findings = [
    ("B10:H10", "All 29 Irrelevant leads from Apr 5–7 were reviewed. Salesperson: Ashima.",             GREY,         "000000"),
    ("B11:H11", "27 leads are correctly categorized — wrong location, accidental clicks, non-admission.", GREY,        "000000"),
    ("B12:H12", "2 leads flagged as MIS-CATEGORIZED with >=25% future admission potential.",             LIGHT_RED,    RED),
    ("B13:H13", "Overall categorization accuracy for the period: 93.1%",                                LIGHT_GREEN,  GREEN),
]
for r, (merge, text, bg, fc) in enumerate(findings, 10):
    ws1.merge_cells(merge)
    first = merge.split(":")[0]
    ws1[first] = f"  {text}"
    apply(ws1[first], fill=bg, font_color=fc, h_align="left", size=11, border=True)
    ws1.row_dimensions[r].height = 22

ws1.row_dimensions[14].height = 18

# Section: Mis-Categorized Detail
ws1.merge_cells("B15:H15")
ws1["B15"] = "  MIS-CATEGORIZED LEADS — ACTION REQUIRED"
apply(ws1["B15"], fill=RED, bold=True, size=12, h_align="left")
ws1.row_dimensions[15].height = 26

# Header row
headers_flag = ["#", "Phone", "Date", "Salesperson", "Potential %", "Reason", "Recommended Action"]
for ci, h in enumerate(headers_flag, 2):
    c = ws1.cell(row=16, column=ci, value=h)
    apply(c, fill=DARK_GREY, bold=True, size=10)
ws1.row_dimensions[16].height = 20

flagged_data = [
    (1, "+918956366022", "05/04/2026", "Ashima", "35–40%",
     "Called back voluntarily asking for Zirakpur branch",
     "Call back – check if distance manageable or future branch planned"),
    (2, "8146265681", "06/04/2026", "Ashima", "25–30%",
     "Specifically asked about Hostel facility from Haryana",
     "Call back – confirm if hostel available; convert to Relevant if yes"),
]
col_widths_flag = [4, 16, 12, 12, 12, 34, 40]
flag_cols = ["B","C","D","E","F","G","H"]
for ci, (col, w) in enumerate(zip(flag_cols, col_widths_flag)):
    ws1.column_dimensions[col].width = w

for ri, row_data in enumerate(flagged_data, 17):
    bg = LIGHT_RED
    for ci, val in enumerate(row_data, 2):
        c = ws1.cell(row=ri, column=ci, value=val)
        apply(c, fill=bg, font_color="000000", h_align="left", size=10, border=True, wrap=True)
        c.font = Font(color="000000", size=10, name="Calibri",
                      bold=(ci == 6))  # bold the potential col
    ws1.row_dimensions[ri].height = 40

ws1.row_dimensions[19].height = 18

# Section: Reason Breakdown
ws1.merge_cells("B20:H20")
ws1["B20"] = "  CATEGORIZATION BREAKDOWN"
apply(ws1["B20"], fill=MID_BLUE, bold=True, size=12, h_align="left")
ws1.row_dimensions[20].height = 26

breakdown = [
    ("Wrong location / too far",               18, "63%"),
    ("Inquired by mistake / accidental click",   5, "17%"),
    ("Non-admission inquiry",                    3, "10%"),
    ("Dead / no contact",                        1,  "3%"),
    ("Affordability + location issue",           1,  "3%"),
    ("Wrong number",                             1,  "3%"),
]
bh = ["B21:D21","E21:F21","G21:H21"]
bh_vals = ["Reason", "Count", "% of Total"]
for merge, val in zip(bh, bh_vals):
    ws1.merge_cells(merge)
    ws1[merge.split(":")[0]] = val
    apply(ws1[merge.split(":")[0]], fill=DARK_GREY, bold=True, size=10)
ws1.row_dimensions[21].height = 20

for ri, (reason, count, pct) in enumerate(breakdown, 22):
    bg = GREY if ri % 2 == 0 else WHITE
    ws1.merge_cells(f"B{ri}:D{ri}")
    ws1[f"B{ri}"] = f"  {reason}"
    apply(ws1[f"B{ri}"], fill=bg, font_color="000000", h_align="left", size=10, border=True)
    ws1.merge_cells(f"E{ri}:F{ri}")
    ws1[f"E{ri}"] = count
    apply(ws1[f"E{ri}"], fill=bg, font_color="000000", size=10, border=True)
    ws1.merge_cells(f"G{ri}:H{ri}")
    ws1[f"G{ri}"] = pct
    apply(ws1[f"G{ri}"], fill=bg, font_color="000000", size=10, border=True)
    ws1.row_dimensions[ri].height = 20


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — ALL LEADS DETAIL
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Leads Detail")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
ws2["A1"] = "ALL IRRELEVANT LEADS — 5th April to 7th April 2026"
apply(ws2["A1"], fill=DARK_BLUE, bold=True, size=14)
ws2.row_dimensions[1].height = 35

ws2.merge_cells("A2:G2")
ws2["A2"] = f"Total: 29 leads reviewed  |  Mis-categorized: 2  |  Accuracy: 93%"
apply(ws2["A2"], fill=MID_BLUE, size=10)
ws2.row_dimensions[2].height = 20

ws2.row_dimensions[3].height = 10

cols2 = ["#", "Date & Time", "Salesperson", "Phone", "Remarks", "Reason for Irrelevant", "Verdict"]
widths2 = [5, 18, 14, 16, 50, 36, 26]
for ci, (h, w) in enumerate(zip(cols2, widths2), 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
    c = ws2.cell(row=4, column=ci, value=h)
    apply(c, fill=DARK_BLUE, bold=True, size=10)
ws2.row_dimensions[4].height = 22

phone_col = unique_headers[3] if len(unique_headers) > 3 else ""

for ri, (_, row) in enumerate(fdf.iterrows(), 5):
    is_flagged = "MIS-CATEGORIZED" in row["Verdict"]
    bg     = LIGHT_RED if is_flagged else (GREY if ri % 2 == 0 else WHITE)
    fc     = RED if is_flagged else "000000"
    values = [
        ri - 4,
        row[col_c],
        row[col_k],
        row[unique_headers[3]] if len(unique_headers) > 3 else "",
        row[col_h].replace("\n", " ").strip(),
        row["Reason"],
        row["Verdict"],
    ]
    for ci, val in enumerate(values, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        apply(c, fill=bg, font_color=fc, h_align="left", size=9,
              border=True, wrap=(ci == 5))
        if is_flagged:
            c.font = Font(color=fc, size=9, bold=True, name="Calibri")
    ws2.row_dimensions[ri].height = 40 if is_flagged else 30


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — FLAGGED LEADS (deep dive)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Flagged Leads")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
ws3["A1"] = "MIS-CATEGORIZED LEADS — DEEP DIVE ANALYSIS"
apply(ws3["A1"], fill=RED, bold=True, size=15)
ws3.row_dimensions[1].height = 38

ws3.merge_cells("A2:F2")
ws3["A2"] = "These leads were marked Irrelevant but show ≥25% potential for future admission."
apply(ws3["A2"], fill=LIGHT_RED, font_color=RED, size=11, h_align="left")
ws3.row_dimensions[2].height = 22
ws3.row_dimensions[3].height = 14

col_widths3 = [22, 22, 22, 22, 22, 22]
col_labels3 = ["A","B","C","D","E","F"]
for col, w in zip(col_labels3, col_widths3):
    ws3.column_dimensions[col].width = w

def flag_block(ws, start_row, lead_num, phone, date, salesperson,
               remarks, reason, potential, action):
    # Title bar
    ws.merge_cells(f"A{start_row}:F{start_row}")
    ws[f"A{start_row}"] = f"  LEAD {lead_num}  |  {phone}  |  {date}  |  Salesperson: {salesperson}"
    apply(ws[f"A{start_row}"], fill=RED, bold=True, size=12, h_align="left")
    ws.row_dimensions[start_row].height = 28

    fields = [
        ("Caller's Remark",    remarks,   LIGHT_RED,    "000000", 60),
        ("Why Flagged",        reason,    LIGHT_ORANGE, "000000", 40),
        ("Admission Potential",potential, LIGHT_GREEN,  GREEN,    22),
        ("Recommended Action", action,    LIGHT_BLUE,   MID_BLUE, 40),
    ]
    for i, (label, value, bg, fc, height) in enumerate(fields):
        r = start_row + 1 + i
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"] = label
        apply(ws[f"A{r}"], fill=DARK_GREY, bold=True, size=10)
        ws.merge_cells(f"C{r}:F{r}")
        ws[f"C{r}"] = value
        apply(ws[f"C{r}"], fill=bg, font_color=fc, h_align="left",
              size=10, border=True, wrap=True)
        ws.row_dimensions[r].height = height

flag_block(
    ws3, 4, 1,
    phone="+918956366022",
    date="05/04/2026",
    salesperson="Ashima",
    remarks="He picked and disconnected and again called me back and said he might had inquired in a hope if we are having another branch near Zirakpur.",
    reason="Parent voluntarily called back after disconnecting — clear sign of genuine interest. He knows the school and was hoping for a Zirakpur branch. This is not an accidental inquiry.",
    potential="35–40% — High potential. If the school has or plans a branch near Zirakpur / Chandigarh corridor, this lead becomes immediately relevant.",
    action="Call back and explain distance options. Ask if they can visit the Patiala campus. If expansion is planned near Zirakpur, add to future pipeline. Do NOT keep as Irrelevant."
)

ws3.row_dimensions[9].height = 18

flag_block(
    ws3, 10, 2,
    phone="8146265681",
    date="06/04/2026",
    salesperson="Ashima",
    remarks="He's from Haryana and asking for Hostel.",
    reason="A parent specifically asking about hostel facility is making a deliberate, informed inquiry — not an accident. He is from Haryana and is open to sending his child to a boarding school, which means distance is not a barrier.",
    potential="25–30% — If the school has hostel facility, this lead is fully relevant. The caller has intent and distance is not an objection since he is seeking boarding.",
    action="Immediately check if school has hostel. If YES — mark as Relevant and schedule a callback. If NO — explain day-boarding options or keep in future pipeline if hostel is planned."
)

ws3.row_dimensions[15].height = 18

# Footer note
ws3.merge_cells("A16:F16")
ws3["A16"] = "Note: Potential % indicates estimated likelihood of admission IF school addresses the specific barrier (location/hostel). Both leads require a callback within 48 hours."
apply(ws3["A16"], fill=LIGHT_BLUE, font_color=MID_BLUE, h_align="left", size=9, wrap=True)
ws3.row_dimensions[16].height = 30


# ══════════════════════════════════════════════════════════════════════════════
#  Save
# ══════════════════════════════════════════════════════════════════════════════
wb.save(REPORT_PATH)
print(f"Report saved: {REPORT_PATH}")
